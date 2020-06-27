from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.utils import get_column_letter
from sys import argv

try:                                               #Tries to open image name without adding 'jpg', with adding 'jpg' or opens default filename
    try:
        im = Image.open(str(argv[1]), 'r')
    except:
        im = Image.open(str(argv[1]) + '.jpg', 'r')
except:
    im = Image.open('testImg.jpg', 'r')
pix_val = list(im.getdata())
pix_val_flat = [x for sets in pix_val for x in sets]

wb = Workbook()
ws = wb.active

def rgb2hex(r,g,b):                                     #Function to convert RGB to HEX value
    return "{:02x}{:02x}{:02x}".format(r,g,b)

def main():
    count = 0
    for x in range(1,im.size[1] + 1):
        print("row no: " + str(x))
        for y in range(1,im.size[0] + 1):
            print("row no: " + str(x) + " | column no: " + str(y))
            cell_color = rgb2hex(pix_val_flat[count], pix_val_flat[count + 1], pix_val_flat[count + 2])
            ws.cell(row=x, column=y).fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type = 'solid')
            count += 3

    for col in range(1,im.size[0] + 1):                     #Sets column width to 3
        ws.column_dimensions[get_column_letter(col)].width = 2

    print("***Worksheet processing done***")

    try:
        wb.save(str(argv[2]) + '.xlsx')                     #Saves workbook into .xlsx format using name as 2nd cmd line arg
    except:
        wb.save('image.xlsx')                               #Saves workbook into .xlsx format using name 'image'

    try:
        print("***Worksheet saving as " + argv[2] + ".xlsx ***")
    except:
        print("***Worksheet saving as image.xlsx ***")

    print("***Save successful***")


if __name__ == "__main__":
    main()